import datetime
import re
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _format_date_value(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        parsed = datetime.datetime.strptime(value, "%Y-%m-%d").date()
        return parsed.strftime("%d.%m.%Y")
    except ValueError:
        return value


def build_schedule_xlsx(
    lessons: Iterable[dict],
    sheet_title: str,
    view: str,
) -> bytes:
    lesson_list = list(lessons)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Расписание"

    if view == "student":
        headers = [
            "Дата",
            "День недели",
            "№ пары",
            "Начало",
            "Конец",
            "Предмет",
            "Тип",
            "Преподаватель",
            "Аудитория",
            "Корпус",
            "Группа",
            "Подгруппа",
        ]
    elif view == "teacher":
        headers = [
            "Дата",
            "День недели",
            "№ пары",
            "Начало",
            "Конец",
            "Предмет",
            "Тип",
            "Аудитория",
            "Корпус",
            "Группа",
            "Подгруппа",
        ]
    else:
        raise ValueError("Unknown view")

    ws.append(headers)
    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for lesson in lesson_list:
        if view == "student":
            ws.append(
                [
                    _format_date_value(lesson.get("date")),
                    lesson.get("day_of_week") or "",
                    lesson.get("lesson_number") or "",
                    lesson.get("start_time") or "",
                    lesson.get("end_time") or "",
                    lesson.get("subject") or "",
                    lesson.get("lesson_type") or "",
                    lesson.get("teacher") or "",
                    lesson.get("room") or "",
                    lesson.get("building") or "",
                    lesson.get("group_name") or "",
                    lesson.get("subgroup_name") or "",
                ]
            )
        else:
            ws.append(
                [
                    _format_date_value(lesson.get("date")),
                    lesson.get("day_of_week") or "",
                    lesson.get("lesson_number") or "",
                    lesson.get("start_time") or "",
                    lesson.get("end_time") or "",
                    lesson.get("subject") or "",
                    lesson.get("lesson_type") or "",
                    lesson.get("room") or "",
                    lesson.get("building") or "",
                    lesson.get("group_name") or "",
                    lesson.get("subgroup_name") or "",
                ]
            )

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


def build_subjects_xlsx(subjects: list[str], sheet_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Предметы"

    ws.append(["№", "Предмет"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["B1"].alignment = Alignment(horizontal="center")

    for i, subject in enumerate(subjects, 1):
        ws.append([i, subject])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A2"

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


def build_safe_filename(name: str, extension: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9А-Яа-яЁё._-]+", "_", name.strip())
    safe = safe.strip("._") or "export"
    return f"{safe}.{extension.lstrip('.')}"
